"""
    FuzzyMatch
    ==========

    This program processes the names and departments of individuals mentioned as
    having a significant impact on students and the course number and title of
    classes students mention as favorites during their time at Tufts.

    Written by:  James Garijo-Garde
            on:  5/2/2019
            for: The Tufts University Office of Institutional Research &
                 Evaluation (OIRE)
"""

import copy
from openpyxl import load_workbook
from fuzzywuzzy import fuzz

# for debugging
import os


# Provides an initial launch message to the user and prompts for the survey mode
def launch():
    print("\n###############################################################################")
    print("### Fuzzy String Matcher for 'Significant Impact' and 'Best Course' Surveys ###")
    print("###############################################################################\n")
    print("Type a two-letter mode code to begin, or visit")
    print("https://github.com/JdoubleG/OIRE-Fuzzy-String for documentation.\n")
    print("'si' = 'Significant Impact'  'bc' = 'Best Course'\n")
    mode = input("Mode code:  ")
    # loop to ensure input is valid
    loop = True
    while loop == True:
        if mode == "si":
            print("\n### Significant Impact ###")
            loop = False
            # Launches the significant impact survey processor
            significantImpact()
        elif mode == "bc":
            print("\n### Best Course ###")
            loop = False
            # Launches the best course survey processor
            bestCourse()
        else:
            mode = input("Input unrecognized. Please re-enter mode code:  ")

def significantImpact():
    print(os.getcwd())
    try:
        wb = load_workbook("si_input.xlsx")
    except:
        print("\nERROR: Input file could not be opened! Ensure that it is correctly named")
        print("'si_input.xlsx' and that it is in the same directory as this script before")
        print("attempting to run it again.")
        exit()
    try:
        ws = wb["RAW"]
    except:
        print("\nERROR: Data sheet could not be opened! Ensure that it is correctly named 'RAW'")
        print("before attempting to run the script again.")
        exit()
    end = ws.max_row
    
    ## fuzzy comparisons of names
    i = 2  # starting index of the new grouping of names
    # 3 = the column with the existing names
    prevName = str(ws.cell(2, 3).value)
    nextName = str(ws.cell(3, 3).value)
    if prevName == None:
        prevName = "-99"
    if nextName == None:
        nextName = "-99"
    # a dictionary of names appearing in the file similar to each other
    names = {}
    if prevName != "-99":
        names.setdefault(prevName, 1)
    row_index = 2
    while row_index <= end:
        if nextName != "-99" or prevName != "-99":
            # 90 = threshold of minimum allowable similarity after passing it
            # into the FuzzyWuzzy token sort algorithm.
            if fuzz.token_sort_ratio(prevName, nextName) >= 90:  # TEST THIS THRESHOLD!
                if nextName in names:
                    # if an instance of the name is already in the dictionary,
                    # update the value
                    names[nextName] = names[nextName] + 1
                else:
                    # otherwise, add the instance of the name to the dictionary
                    names.setdefault(nextName, 1)
            else:
                # find the key with the max value (the one appearing most often
                # in the data)
                keys = list(names.keys())
                if len(keys) > 1:
                    topScore = 0
                    bestName = ""
                    for j in keys:
                        if names[j] > topScore:
                            topScore = names[j]
                            bestName = j
                else:
                    bestName = prevName
                # for all elements with similar names, give them the same name
                for j in range(i, row_index + 1):
                    # 4 = the column with the new names
                    ws.cell(j, 4).value = bestName
                # update the starting index of the new grouping of names
                i = row_index + 1
                # reset the names dictionary
                names = {}
                if prevName != "-99":
                    names.setdefault(prevName, 1)
        else:
            i = row_index + 1
        # update the row index
        row_index = row_index + 1
        prevName = str(ws.cell(row_index, 3).value)
        nextName = str(ws.cell(row_index + 1, 3).value)
        if prevName == None:
            prevName = "-99"
        if nextName == None:
            nextName = "-99"
    
    ## non-fuzzy comparison of department/subject
    i = 2  # starting index of the new grouping of subjects
    prevSubject = str(ws.cell(2, 5).value)
    nextSubject = str(ws.cell(3, 5).value)
    if prevSubject == None:
        prevSubject = "-99"
    if nextSubject == None:
        nextSubject = "-99"
    # a dictionary of subjects appearing in the file similar to each other
    subjects = {prevSubject: 1}
    row_index = 2
    while row_index <= end:
        # The subject loop adds a check to make sure it is comparing the same
        # person for whom the subject is associated.
        if ws.cell(row_index, 4).value == ws.cell(row_index + 1, 4).value:
            if nextSubject in subjects:
                # if an instance of the subject is already in the
                # dictionary, update the value
                subjects[nextSubject] = subjects[nextSubject] + 1
            else:
                # otherwise, add the instance of the subject to the
                # dictionary
                subjects.setdefault(nextSubject, 1)
        else:
            # find the key with the max value (the on appearing most often in
            # the data)
            keys = list(subjects.keys())
            if len(keys) > 1:
                topScore = 0
                bestSubject = ""
                for j in keys:
                    if subjects[j] > topScore:
                        topScore = subjects[j]
                        bestSubject = j
            else:
                bestSubject = prevSubject
            # for all elements with similar subjects, give them the same
            # subject
            for j in range(i, row_index + 1):
                # 6 = the column with the new subjects
                ws.cell(j, 6).value = str(bestSubject)
            # update the starting index of the new grouping of subjects
            i = row_index + 1
            # reset the subjects dictionary
            subjects = {nextSubject: 1}
        # update the row index
        row_index = row_index + 1
        prevSubject = str(ws.cell(row_index, 5).value)
        nextSubject = str(ws.cell(row_index + 1, 5).value)
        if prevSubject == None:
            prevSubject = "-99"
        if nextSubject == None:
            nextSubject = "-99"

    ws.title = "OUTPUT"
    try:
        wb.save("si_output.xlsx")
    except:
        print("\nERROR: Data sheet could not be saved!")
        exit()
    print("\n\nSurvey fuzzy string processing complete! File saved as 'si_output.xlsx\n")


def bestCourse():
    try:
        wb = load_workbook("bc_input.xlsx")
    except:
        print("\nERROR: Input file could not be opened! Ensure that it is correctly named")
        print("'bc_input.xlsx' and that it is in the same directory as this script before")
        print("attempting to run it again.")
        exit()
    try:
        ws = wb["RAW"]
    except:
        print("\nERROR: Data sheet could not be opened! Ensure that it is correctly named 'RAW'")
        print("before attempting to run the script again.")
        exit()
    end = ws.max_row
    # do fuzzy string stuff

    ## Format the course number in a standard way
    for i in range(2, end + 1):
        # 4 = the column with the existing course numbers
        courseNum = str(ws.cell(i, 4).value)
        if courseNum != "-99":
            # capitalize all the letters in the course number
            courseNum = courseNum.upper()
            for j in range(len(courseNum)):
                if courseNum[j].isalnum() == False and courseNum[j + 1].isdigit():
                    # if a character in the course number is neither a letter
                    # nor a number, change it to a space
                    courseNum = courseNum[:j] + ' ' + courseNum[(j + 1):]
                elif courseNum[j].isdigit() and j == 0:
                    # if the course number does not have the subject code at the
                    # beginning, get the subject code from the subject column
                    num = ""
                    extra = False
                    for k in range(len(courseNum)):
                        if courseNum[k].isdigit():
                            num = num + courseNum[k]
                        elif k < len(courseNum) - 1:
                            extra = True
                            break
                    while len(num) < 4:
                        num = "0" + num
                    subject = str(ws.cell(i, 3).value)
                    # if the subject field is empty, don't do anything
                    if subject != "-99":
                        start = 0
                        # parse out the subject code from the value in the
                        # subject column
                        for k in range(len(subject)):
                            if subject[k] == '(':
                                start = k + 1
                                break
                        subject = subject[start:(len(subject) - 1)] + ' '
                    newCourseNum = copy.deepcopy(courseNum)
                    # insert a space if the course number goes straight from
                    # letters to numbers
                    newCourseNum = subject + num
                    if extra:
                        newCourseNum = newCourseNum + courseNum[(k):]
                    courseNum = newCourseNum
                    break
                elif courseNum[j].isdigit():
                    # make sure the appropriate number of leading zeros are in
                    # the course number
                    num = ""
                    extra = False
                    for k in range(j, len(courseNum)):
                        if courseNum[k].isdigit():
                            num = num + courseNum[k]
                        elif k < len(courseNum) - 1:
                            extra = True
                            break
                    while len(num) < 4:
                        num = "0" + num
                    newCourseNum = copy.deepcopy(courseNum)
                    # insert a space if the course number goes straight from
                    # letters to numbers
                    if courseNum[j-1].isalpha():
                        newCourseNum = courseNum[:j] + " " + num
                    else:
                        newCourseNum = courseNum[:j] + num
                    if extra:
                        newCourseNum = newCourseNum + courseNum[(k):]
                    courseNum = newCourseNum
                    break
        # 5 = the column with the new course numbers
        ws.cell(i, 5).value = courseNum

    try:
        wb.save("bc_output.xlsx")
    except:
        print("\nERROR: Data sheet could not be saved!")
        exit()
    print("\n\nPhase 1 complete!\n")
    print("Please open 'bc_output.xlsx' and, in the 'RAW' sheet, sort the new course")
    print("numbers alphabetically before continuing.\n")
    input("Press the enter key when you are ready to continue.")

    try:
        wb = load_workbook("bc_output.xlsx")
    except:
        print("\nERROR: Input file could not be opened! Ensure that it is correctly named")
        print("'bc_output.xlsx' and that it is in the same directory as this script before")
        print("attempting to run it again.")
        exit()
    try:
        ws = wb["RAW"]
    except:
        print("\nERROR: Data sheet could not be opened! Ensure that it is correctly named 'RAW'")
        print("before attempting to run the script again.")
        exit()
    
    ## non-fuzzy comparison of course titles
    i = 2  # starting index of the new grouping of course titles
    prevTitle = str(ws.cell(2, 6).value)
    nextTitle = str(ws.cell(3, 6).value)
    if prevTitle == None:
        prevTitle = "-99"
    if nextTitle == None:
        nextTitle = "-99"
    # a dictionary of course titles appearing in the file similar to each other
    titles = {}
    if prevTitle != "-99":
        titles.setdefault(prevTitle, 1)
    row_index = 2
    while row_index <= end:
        if str(ws.cell(row_index, 5).value) == "-99":
            # if the course number is -99, put the user-entered course title in
            # the column of new course titles.
            ws.cell(row_index, 7).value = ws.cell(row_index, 6).value
            # update the starting index of the new grouping of course titles
            i = row_index + 1
        elif nextTitle != "-99" or prevTitle != "-99":  # and ws.cell(row_index + 1, 5).value != "-99"
            # The course title loop adds a check to make sure it is comparing the
            # same course number for which the course title is associated.
            if str(ws.cell(row_index, 5).value) == str(ws.cell(row_index + 1, 5).value):
                if nextTitle in titles:
                    # if an instance of the title is already in the dictionary,
                    # update the value
                    titles[nextTitle] = titles[nextTitle] + 1
                else:
                    # otherwise, add the instance of the course title to the
                    # dictionary
                    titles.setdefault(nextTitle, 1)
            else:
                # find the key with the max value (the on appearing most often in
                # the data)
                keys = list(titles.keys())
                if len(keys) > 1:
                    topScore = 0
                    bestTitle = ""
                    for j in keys:
                        if titles[j] > topScore and j != "-99":
                            topScore = titles[j]
                            bestTitle = j
                else:
                    bestTitle = prevTitle
                # for all elements with similar titles, give them the same title
                for j in range(i, row_index + 1):
                    # 7 = the column with the new course titles
                    ws.cell(j, 7).value = str(bestTitle)
                # update the starting index of the new grouping of course titles
                i = row_index + 1
                # reset the titles dictionary
                titles = {}
                if nextTitle != "-99":
                    titles.setdefault(nextTitle, 1)
        # update the row index
        row_index = row_index + 1
        prevTitle = str(ws.cell(row_index, 6).value)
        nextTitle = str(ws.cell(row_index + 1, 6).value)
        if prevTitle == None:
            prevTitle = "-99"
        if nextTitle == None:
            nextTitle = "-99"

    try:
        wb.save("bc_output.xlsx")
    except:
        print("\nERROR: Data sheet could not be saved!")
        exit()
    print("\n\nPhase 2 complete!\n")
    print("Please open 'bc_output.xlsx' and, in the 'RAW' sheet, sort the new course titles")
    print("alphabetically before continuing.\n")
    input("Press the enter key when you are ready to continue.")

    try:
        wb = load_workbook("bc_output.xlsx")
    except:
        print("\nERROR: Input file could not be opened! Ensure that it is correctly named")
        print("'bc_output.xlsx' and that it is in the same directory as this script before")
        print("attempting to run it again.")
        exit()
    try:
        ws = wb["RAW"]
    except:
        print("\nERROR: Data sheet could not be opened! Ensure that it is correctly named 'RAW'")
        print("before attempting to run the script again.")
        exit()

    ## fuzzy comparison of course titles
    i = 2  # starting index of the new grouping of titles
    # 6 = the column with the titles
    prevTitle = str(ws.cell(2, 6).value)
    nextTitle = str(ws.cell(3, 6).value)
    if prevTitle == None:
        prevTitle = "-99"
    if nextTitle == None:
        nextTitle = "-99"
    # a dictionary of titles appearing in the file similar to each other
    titles = {}
    if prevTitle != "-99":
        titles.setdefault(prevTitle, 1)
    row_index = 2
    while row_index <= end:
        if nextTitle != "-99" or prevTitle != "-99":
            # 90 = threshold of minimum allowable similarity after passing it
            # into the FuzzyWuzzy token sort algorithm.
            if fuzz.token_sort_ratio(prevTitle, nextTitle) >= 90:  # TEST THIS THRESHOLD!
                if nextTitle in titles:
                    # if an instance of the title is already in the dictionary,
                    # update the value
                    titles[nextTitle] = titles[nextTitle] + 1
                else:
                    # otherwise, add the instance of the title to the dictionary
                    titles.setdefault(nextTitle, 1)
            else:
                # find the key with the max value (the one appearing most often
                # in the data)
                keys = list(titles.keys())
                if len(keys) > 1:
                    topScore = 0
                    bestTitle = ""
                    for j in keys:
                        if titles[j] > topScore:
                            topScore = titles[j]
                            bestTitle = j
                else:
                    bestTitle = prevTitle
                # for all elements with similar titles, give them the same title
                for j in range(i, row_index + 1):
                    # 4 = the column with the new title
                    ws.cell(j, 7).value = bestTitle
                # update the starting index of the new grouping of names
                i = row_index + 1
                # reset the titles dictionary
                titles = {}
                if prevTitle != "-99":
                    titles.setdefault(prevTitle, 1)
        else:
            i = row_index + 1
        # update the row index
        row_index = row_index + 1
        prevTitle = str(ws.cell(row_index, 6).value)
        nextTitle = str(ws.cell(row_index + 1, 6).value)
        if prevTitle == None:
            prevTitle = "-99"
        if nextTitle == None:
            nextTitle = "-99"

    ## Add course number and course title to rows with incomplete information
    for i in range(2, end + 1):
        if ws.cell(i, 5).value == "-99":
            if ws.cell(i, 7).value == ws.cell(i - 1, 7).value:
                ws.cell(i, 5).value = ws.cell(i - 1, 5).value
            elif ws.cell(i, 7).value == ws.cell(i + 1, 7).value:
                ws.cell(i, 5).value = ws.cell(i + 1, 5).value 
        elif ws.cell(i, 7).value == "-99":
            if ws.cell(i, 5).value == ws.cell(i - 1, 5).value:
                ws.cell(i, 7).value = ws.cell(i - 1, 7).value
            elif ws.cell(i, 5).value == ws.cell(i + 1, 5).value:
                ws.cell(i, 7).value = ws.cell(i + 1, 7).value

    ws.title = "OUTPUT"
    try:
        wb.save("bc_output.xlsx")
    except:
        print("\nERROR: Data sheet could not be saved!")
        exit()
    print("\n\nSurvey fuzzy string processing complete! File saved as 'bc_output.xlsx\n")


#########  MAIN  #########
if __name__ == "__main__":
    launch()
    input("Press the enter key to exit.")